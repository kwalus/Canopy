"""
Read-only file preview helpers for text and spreadsheet attachments.

This module intentionally does not execute embedded spreadsheet code or VBA.
It extracts a bounded preview suitable for inline Canopy rendering.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency declared in pyproject
    load_workbook = None


SPREADSHEET_MIME_TYPES = {
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroenabled.12",
}

SPREADSHEET_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
MARKDOWN_EXTENSIONS = {".md", ".markdown"}
DOCUMENT_PREVIEW_EXTENSIONS = {
    ".docx",
    ".docm",
    ".dotx",
    ".pptx",
    ".pptm",
    ".ppsx",
    ".potx",
    ".rtf",
    ".odt",
    ".odp",
}
DOCUMENT_PREVIEW_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-word.document.macroenabled.12",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.template",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "application/vnd.ms-powerpoint.presentation.macroenabled.12",
    "application/vnd.openxmlformats-officedocument.presentationml.slideshow",
    "application/vnd.openxmlformats-officedocument.presentationml.template",
    "application/rtf",
    "text/rtf",
    "application/vnd.oasis.opendocument.text",
    "application/vnd.oasis.opendocument.presentation",
}
TEXT_PREVIEW_EXTENSIONS = {
    ".md",
    ".markdown",
    ".txt",
    ".log",
    ".json",
    ".py",
    ".js",
    ".ts",
    ".csv",
    ".tsv",
    ".yaml",
    ".yml",
    ".xml",
    ".tex",
    ".html",
    ".css",
    ".sh",
    ".bat",
    ".cfg",
    ".ini",
    ".toml",
}
TEXT_PREVIEW_MIME_PREFIXES = ("text/",)
TEXT_PREVIEW_MIME_TYPES = {
    "application/json",
    "application/xml",
    "application/x-yaml",
    "application/javascript",
    "application/typescript",
    "text/x-tex",
    "application/x-latex",
}
CANOPY_MODULE_SUFFIXES = (".canopy-module.html", ".canopy-module.htm")

MAX_TEXT_PREVIEW_BYTES = 512 * 1024
MAX_TEXT_PREVIEW_CHARS = 50_000
MAX_DOCUMENT_PREVIEW_BYTES = 16 * 1024 * 1024
MAX_DOCUMENT_PREVIEW_CHARS = 40_000
MAX_PRESENTATION_SLIDES = 20
MAX_SPREADSHEET_PREVIEW_BYTES = 12 * 1024 * 1024
MAX_SHEETS = 3
MAX_ROWS = 60
MAX_COLS = 14
MAX_CELL_CHARS = 160


def _file_extension(filename: str | None) -> str:
    return Path(filename or "").suffix.lower()


def is_canopy_module_bundle(filename: str | None, content_type: str | None) -> bool:
    lower_name = str(filename or "").strip().lower()
    lower_type = str(content_type or "").strip().lower()
    return lower_type == "text/html" and any(lower_name.endswith(suffix) for suffix in CANOPY_MODULE_SUFFIXES)


def is_markdown_previewable(filename: str | None, content_type: str | None) -> bool:
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    return ext in MARKDOWN_EXTENSIONS or ctype in {"text/markdown", "text/x-markdown"}


def is_spreadsheet_previewable(filename: str | None, content_type: str | None) -> bool:
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    return ext in SPREADSHEET_EXTENSIONS or ctype in SPREADSHEET_MIME_TYPES


def is_document_previewable(filename: str | None, content_type: str | None) -> bool:
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    return ext in DOCUMENT_PREVIEW_EXTENSIONS or ctype in DOCUMENT_PREVIEW_MIME_TYPES


def is_text_previewable(filename: str | None, content_type: str | None) -> bool:
    if is_canopy_module_bundle(filename, content_type):
        return False
    if is_spreadsheet_previewable(filename, content_type):
        return False
    if is_document_previewable(filename, content_type):
        return False
    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    if ext in TEXT_PREVIEW_EXTENSIONS:
        return True
    if ctype in TEXT_PREVIEW_MIME_TYPES:
        return True
    return any(ctype.startswith(prefix) for prefix in TEXT_PREVIEW_MIME_PREFIXES)


def _decode_text_bytes(file_data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "utf-16"):
        try:
            return file_data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_data.decode("utf-8", errors="replace")


def _truncate_text(value: str, limit: int) -> tuple[str, bool]:
    if len(value) <= limit:
        return value, False
    return value[:limit], True


def _normalize_preview_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n{4,}", "\n\n\n", value)
    return value.strip()


def _xml_text_chunks(xml_bytes: bytes, text_tags: set[str], paragraph_tags: set[str] | None = None) -> list[str]:
    paragraph_tags = paragraph_tags or set()
    chunks: list[str] = []
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return chunks
    for elem in root.iter():
        tag = str(elem.tag or "").rsplit("}", 1)[-1]
        if tag in text_tags and elem.text:
            chunks.append(elem.text)
        elif tag in paragraph_tags and chunks and chunks[-1] != "\n":
            chunks.append("\n")
    return chunks


def _extract_docx_text(file_data: bytes) -> str:
    parts: list[str] = []
    with zipfile.ZipFile(io.BytesIO(file_data)) as archive:
        names = archive.namelist()
        ordered_names = ["word/document.xml"]
        ordered_names.extend(name for name in names if name.startswith("word/header") and name.endswith(".xml"))
        ordered_names.extend(name for name in names if name.startswith("word/footer") and name.endswith(".xml"))
        for name in ordered_names:
            if name not in names:
                continue
            chunks = _xml_text_chunks(
                archive.read(name),
                text_tags={"t"},
                paragraph_tags={"p", "br"},
            )
            text = _normalize_preview_text("".join(chunks))
            if text:
                parts.append(text)
    return "\n\n".join(parts)


def _slide_sort_key(name: str) -> tuple[int, str]:
    match = re.search(r"slide(\d+)\.xml$", name)
    return (int(match.group(1)) if match else 0, name)


def _extract_pptx_text(file_data: bytes) -> tuple[str, bool]:
    sections: list[str] = []
    truncated = False
    with zipfile.ZipFile(io.BytesIO(file_data)) as archive:
        slide_names = sorted(
            [name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")],
            key=_slide_sort_key,
        )
        if len(slide_names) > MAX_PRESENTATION_SLIDES:
            truncated = True
        for index, name in enumerate(slide_names[:MAX_PRESENTATION_SLIDES], start=1):
            chunks = _xml_text_chunks(archive.read(name), text_tags={"t"})
            text = _normalize_preview_text("\n".join(chunk.strip() for chunk in chunks if chunk.strip()))
            if text:
                sections.append(f"Slide {index}\n{text}")
            else:
                sections.append(f"Slide {index}\n(no extractable text)")
    return "\n\n".join(sections), truncated


def _extract_opendocument_text(file_data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(file_data)) as archive:
        if "content.xml" not in archive.namelist():
            return ""
        chunks = _xml_text_chunks(
            archive.read("content.xml"),
            text_tags={"h", "p", "span", "a"},
            paragraph_tags={"h", "p"},
        )
    return _normalize_preview_text(" ".join(chunk.strip() for chunk in chunks if chunk.strip()))


def _extract_rtf_text(file_data: bytes) -> str:
    raw = _decode_text_bytes(file_data[:MAX_TEXT_PREVIEW_BYTES])
    raw = re.sub(r"\\'[0-9a-fA-F]{2}", " ", raw)
    raw = re.sub(r"\\par[d]?", "\n", raw)
    raw = re.sub(r"\\tab", "\t", raw)
    raw = re.sub(r"{\\\*[^{}]*}", " ", raw)
    raw = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", raw)
    raw = raw.replace("{", " ").replace("}", " ").replace("\\", "")
    raw = re.sub(r"[ \t]{2,}", " ", raw)
    return _normalize_preview_text(raw)


def _serialize_cell(value: Any) -> dict[str, Any]:
    if value is None:
        return {"display": "", "kind": "empty"}
    if isinstance(value, bool):
        return {"display": "TRUE" if value else "FALSE", "kind": "boolean"}
    if isinstance(value, int):
        return {"display": str(value), "kind": "number"}
    if isinstance(value, float):
        if value.is_integer():
            return {"display": str(int(value)), "kind": "number"}
        return {"display": format(value, ".15g"), "kind": "number"}
    if isinstance(value, datetime):
        return {"display": value.isoformat(sep=" ", timespec="seconds"), "kind": "datetime"}
    if isinstance(value, date):
        return {"display": value.isoformat(), "kind": "date"}
    if isinstance(value, time):
        return {"display": value.isoformat(timespec="seconds"), "kind": "time"}

    display = str(value)
    truncated_display, truncated = _truncate_text(display, MAX_CELL_CHARS)
    cell = {"display": truncated_display, "kind": "text"}
    if truncated:
        cell["truncated"] = True
        cell["full_length"] = len(display)
    return cell


def _build_text_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    preview_bytes = file_data[:MAX_TEXT_PREVIEW_BYTES]
    text = _decode_text_bytes(preview_bytes)
    text, text_truncated = _truncate_text(text, MAX_TEXT_PREVIEW_CHARS)
    kind = "markdown" if is_markdown_previewable(filename, content_type) else "text"
    return {
        "previewable": True,
        "kind": kind,
        "text": text,
        "truncated": len(file_data) > MAX_TEXT_PREVIEW_BYTES or text_truncated,
        "limits": {
            "max_bytes": MAX_TEXT_PREVIEW_BYTES,
            "max_chars": MAX_TEXT_PREVIEW_CHARS,
        },
    }


def _build_document_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    if len(file_data) > MAX_DOCUMENT_PREVIEW_BYTES:
        return {
            "previewable": False,
            "kind": "document",
            "error": (
                f"Document preview is limited to {MAX_DOCUMENT_PREVIEW_BYTES // (1024 * 1024)} MB. "
                "Download the file to inspect the full document."
            ),
        }

    ext = _file_extension(filename)
    ctype = str(content_type or "").lower()
    truncated = False
    try:
        if ext in {".docx", ".docm", ".dotx"} or ctype in {
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.ms-word.document.macroenabled.12",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.template",
        }:
            text = _extract_docx_text(file_data)
        elif ext in {".pptx", ".pptm", ".ppsx", ".potx"} or ctype in {
            "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            "application/vnd.ms-powerpoint.presentation.macroenabled.12",
            "application/vnd.openxmlformats-officedocument.presentationml.slideshow",
            "application/vnd.openxmlformats-officedocument.presentationml.template",
        }:
            text, truncated = _extract_pptx_text(file_data)
        elif ext in {".odt", ".odp"} or ctype in {
            "application/vnd.oasis.opendocument.text",
            "application/vnd.oasis.opendocument.presentation",
        }:
            text = _extract_opendocument_text(file_data)
        elif ext == ".rtf" or ctype in {"application/rtf", "text/rtf"}:
            text = _extract_rtf_text(file_data)
        else:
            text = ""
    except zipfile.BadZipFile:
        return {
            "previewable": False,
            "kind": "document",
            "error": "Document preview is unavailable because the file container is not readable.",
        }
    except Exception:
        return {
            "previewable": False,
            "kind": "document",
            "error": "Document preview could not extract readable text from this file.",
        }

    text = _normalize_preview_text(text)
    if not text:
        return {
            "previewable": False,
            "kind": "document",
            "error": "No readable text was found for inline preview. Download the file to inspect it.",
        }

    text, text_truncated = _truncate_text(text, MAX_DOCUMENT_PREVIEW_CHARS)
    macro_enabled = ext in {".docm", ".pptm"} or ctype in {
        "application/vnd.ms-word.document.macroenabled.12",
        "application/vnd.ms-powerpoint.presentation.macroenabled.12",
    }
    return {
        "previewable": True,
        "kind": "document",
        "document_format": ext.lstrip(".") or "document",
        "macro_enabled": macro_enabled,
        "text": text,
        "truncated": truncated or text_truncated,
        "warning": (
            "Document preview is read-only. Canopy never executes Office macros or embedded active content."
            if macro_enabled else None
        ),
        "limits": {
            "max_bytes": MAX_DOCUMENT_PREVIEW_BYTES,
            "max_chars": MAX_DOCUMENT_PREVIEW_CHARS,
            "max_presentation_slides": MAX_PRESENTATION_SLIDES,
        },
    }


def _csv_rows(text: str, filename: str) -> tuple[list[list[dict[str, Any]]], int, int]:
    sample = text[:8192]
    delimiter = "\t" if _file_extension(filename) == ".tsv" else ","
    try:
        sniffed = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = sniffed.delimiter
    except Exception:
        pass

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[dict[str, Any]]] = []
    total_rows = 0
    total_cols = 0
    for raw_row in reader:
        total_rows += 1
        total_cols = max(total_cols, len(raw_row))
        if len(rows) >= MAX_ROWS:
            continue
        cooked = [_serialize_cell(value) for value in raw_row[:MAX_COLS]]
        while cooked and cooked[-1]["kind"] == "empty":
            cooked.pop()
        rows.append(cooked)
    return rows, total_rows, total_cols


def _build_csv_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    preview_bytes = file_data[:MAX_TEXT_PREVIEW_BYTES]
    text = _decode_text_bytes(preview_bytes)
    rows, total_rows, total_cols = _csv_rows(text, filename)
    return {
        "previewable": True,
        "kind": "spreadsheet",
        "macro_enabled": False,
        "sheets": [
            {
                "name": Path(filename or "Sheet1").stem or "Sheet1",
                "rows": rows,
                "row_count": total_rows,
                "col_count": total_cols,
                "preview_row_count": len(rows),
                "preview_col_count": min(MAX_COLS, max((len(row) for row in rows), default=0)),
                "truncated_rows": total_rows > MAX_ROWS,
                "truncated_cols": total_cols > MAX_COLS,
            }
        ],
        "sheet_count": 1,
        "truncated": len(file_data) > MAX_TEXT_PREVIEW_BYTES or total_rows > MAX_ROWS or total_cols > MAX_COLS,
        "limits": {
            "max_bytes": MAX_TEXT_PREVIEW_BYTES,
            "max_sheets": 1,
            "max_rows": MAX_ROWS,
            "max_cols": MAX_COLS,
        },
    }


def _build_workbook_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    if load_workbook is None:
        return {
            "previewable": False,
            "kind": "spreadsheet",
            "error": "Spreadsheet preview dependency is unavailable on this Canopy instance.",
        }

    if len(file_data) > MAX_SPREADSHEET_PREVIEW_BYTES:
        return {
            "previewable": False,
            "kind": "spreadsheet",
            "error": (
                f"Spreadsheet preview is limited to {MAX_SPREADSHEET_PREVIEW_BYTES // (1024 * 1024)} MB. "
                "Download the file to inspect the full workbook."
            ),
        }

    workbook = load_workbook(
        io.BytesIO(file_data),
        read_only=True,
        data_only=True,
        keep_vba=False,
    )
    sheets: list[dict[str, Any]] = []
    total_sheet_count = len(workbook.worksheets)

    for worksheet in workbook.worksheets[:MAX_SHEETS]:
        rows: list[list[dict[str, Any]]] = []
        preview_col_count = 0
        for raw_row in worksheet.iter_rows(min_row=1, max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True):
            cooked = [_serialize_cell(value) for value in raw_row]
            while cooked and cooked[-1]["kind"] == "empty":
                cooked.pop()
            rows.append(cooked)
            preview_col_count = max(preview_col_count, len(cooked))

        row_count = int(getattr(worksheet, "max_row", 0) or 0)
        col_count = int(getattr(worksheet, "max_column", 0) or 0)
        sheets.append(
            {
                "name": worksheet.title,
                "rows": rows,
                "row_count": row_count,
                "col_count": col_count,
                "preview_row_count": len(rows),
                "preview_col_count": min(MAX_COLS, preview_col_count),
                "truncated_rows": row_count > MAX_ROWS,
                "truncated_cols": col_count > MAX_COLS,
            }
        )

    try:
        workbook.close()
    except Exception:
        pass

    macro_enabled = _file_extension(filename) == ".xlsm" or str(content_type or "").lower() == (
        "application/vnd.ms-excel.sheet.macroenabled.12"
    )
    return {
        "previewable": True,
        "kind": "spreadsheet",
        "macro_enabled": macro_enabled,
        "sheets": sheets,
        "sheet_count": total_sheet_count,
        "truncated": total_sheet_count > MAX_SHEETS or any(sheet["truncated_rows"] or sheet["truncated_cols"] for sheet in sheets),
        "warning": (
            "Workbook preview is read-only. Canopy never executes spreadsheet macros or VBA."
            if macro_enabled else None
        ),
        "limits": {
            "max_bytes": MAX_SPREADSHEET_PREVIEW_BYTES,
            "max_sheets": MAX_SHEETS,
            "max_rows": MAX_ROWS,
            "max_cols": MAX_COLS,
        },
    }


def build_file_preview(file_data: bytes, filename: str, content_type: str) -> dict[str, Any]:
    if is_canopy_module_bundle(filename, content_type):
        return {
            "previewable": False,
            "kind": "module",
            "error": "Canopy Module bundles open in the deck, not the generic file preview.",
        }
    if is_spreadsheet_previewable(filename, content_type):
        if _file_extension(filename) in {".csv", ".tsv"} or str(content_type or "").lower() == "text/csv":
            return _build_csv_preview(file_data, filename, content_type)
        return _build_workbook_preview(file_data, filename, content_type)
    if is_document_previewable(filename, content_type):
        return _build_document_preview(file_data, filename, content_type)
    if is_text_previewable(filename, content_type):
        return _build_text_preview(file_data, filename, content_type)
    return {
        "previewable": False,
        "kind": "unsupported",
        "error": "Inline preview is not available for this file type.",
    }
